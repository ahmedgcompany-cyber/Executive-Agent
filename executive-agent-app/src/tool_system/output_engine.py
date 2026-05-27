"""
Output Intelligence Engine
==========================
Decides the BEST output format for any task result and exports it externally.

Format selection (goal keywords → format):
  leads / companies / table / outreach  → Excel  (.xlsx)
  json / api / dataset / config         → JSON   (.json)
  report / analysis / audit / writeup   → PDF    (.pdf)   via reportlab
  dashboard / chart / html / visual     → HTML   (.html)
  code / script / python / .py          → Source (.py)
  package / multiple / zip              → ZIP    (.zip)
  default                               → Text   (.txt)

Every export lands on Desktop (OneDrive Desktop as fallback).
Missing dependencies are auto-installed.
"""

from __future__ import annotations

import datetime
import json
import os
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Optional


# ---------------------------------------------------------------------------
# Keyword → format decision table
# ---------------------------------------------------------------------------

_FORMAT_RULES: list[tuple[list[str], str]] = [
    (
        ["lead", "leads", "lead gen", "lead generation", "prospect", "prospects",
         "company list", "companies", "business list", "businesses",
         "outreach list", "contact list", "crm export", "sales list",
         "find companies", "find businesses", "find leads"],
        "excel",
    ),
    (
        ["json", "api response", "api data", "structured data",
         "config", "configuration", "schema", "dataset", "records"],
        "json",
    ),
    (
        ["report", "analysis", "research report", "audit", "pdf",
         "write up", "writeup", "summary report", "findings"],
        "pdf",
    ),
    (
        ["dashboard", "chart", "visualization", "html page",
         "web page", "webpage", "landing page", "interactive"],
        "html",
    ),
    (
        ["write code", "generate code", "python script", "automation script",
         ".py file", "create a script", "build a script"],
        "code",
    ),
    (
        ["zip", "package", "multiple files", "folder", "archive"],
        "zip",
    ),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _desktop() -> Path:
    """Return local Desktop directory (NOT OneDrive)."""
    # Prefer real Desktop path from registry to avoid OneDrive redirect
    desktop = Path.home() / "Desktop"
    try:
        import winreg
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders",
        )
        real_desktop = winreg.QueryValueEx(key, "Desktop")[0]
        if real_desktop:
            desktop = Path(real_desktop)
    except Exception:
        pass
    if desktop.exists():
        return desktop
    # Fallbacks (never OneDrive)
    for c in [
        Path.home() / "Desktop",
        Path.home() / "Documents",
    ]:
        if c.exists():
            return c
    p = Path.home() / "Desktop"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _slug(goal: str, maxlen: int = 40) -> str:
    return re.sub(r"[^\w]+", "_", goal[:maxlen]).strip("_")


def _ts() -> str:
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


def _pip(package: str) -> bool:
    import subprocess
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", package, "-q"],
        capture_output=True, timeout=120,
    )
    return r.returncode == 0


# ---------------------------------------------------------------------------
# Output Intelligence Engine
# ---------------------------------------------------------------------------

class OutputEngine:
    """
    Determines the best output format for a task result and exports the file.

    Usage:
        engine = OutputEngine()
        path = engine.auto_export(goal, result)   # returns saved path or ""
    """

    def decide_format(self, goal: str, result: dict) -> str:
        """
        Analyse goal + result and pick the best export format.

        Returns one of: 'excel', 'json', 'pdf', 'html', 'code', 'zip', 'txt'
        """
        goal_lower = goal.lower()
        summary_lower = (result.get("summary") or "").lower()[:500]
        combined = goal_lower + " " + summary_lower

        for keywords, fmt in _FORMAT_RULES:
            if any(kw in combined for kw in keywords):
                return fmt

        # If result has structured business data → Excel
        if result.get("businesses") or result.get("leads"):
            return "excel"

        # If result is a code block → code file
        summary = result.get("summary") or result.get("content") or ""
        if re.search(r"```(?:python|py|javascript|js|bash|sh)", summary, re.IGNORECASE):
            return "code"

        return "txt"

    def auto_export(
        self,
        goal: str,
        result: dict,
        force_format: Optional[str] = None,
    ) -> str:
        """
        Choose best format and export the result to Desktop.

        Returns the saved file path, or "" on failure.
        """
        # Conversational fast-path replies opt out of file export.
        if isinstance(result, dict) and result.get("skip_export"):
            return ""
        fmt = force_format or self.decide_format(goal, result)

        exporters = {
            "excel": self._export_excel,
            "json":  self._export_json,
            "pdf":   self._export_pdf,
            "html":  self._export_html,
            "code":  self._export_code,
            "zip":   self._export_zip,
            "txt":   self._export_txt,
        }

        exporter = exporters.get(fmt, self._export_txt)
        try:
            path = exporter(goal, result)
            return path or ""
        except Exception:
            # Fallback to txt if chosen format fails
            try:
                return self._export_txt(goal, result) or ""
            except Exception:
                return ""

    # ------------------------------------------------------------------
    # Format exporters
    # ------------------------------------------------------------------

    # ── Excel ──────────────────────────────────────────────────────────
    def _export_excel(self, goal: str, result: dict) -> Optional[str]:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            if not _pip("openpyxl"):
                return self._export_txt(goal, result)
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        businesses: list[dict] = result.get("businesses") or result.get("leads") or []
        summary: str = result.get("summary") or result.get("content") or ""
        ts = _ts()
        path = _desktop() / f"MegaV_{_slug(goal)}_{ts}.xlsx"

        wb = openpyxl.Workbook()

        # ── Shared styles ───────────────────────────────────────────
        _hdr_font = Font(bold=True, color="FFFFFF", size=11)
        _hdr_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        _alt_fill = PatternFill(fill_type="solid", fgColor="D6E4F0")
        _ctr      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _wrap     = Alignment(vertical="top", wrap_text=True)
        _thin     = Side(style="thin", color="AAAAAA")
        _border   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        # ── Sheet 1: Leads / Data ───────────────────────────────────
        ws = wb.active
        ws.title = "Leads"
        ws.freeze_panes = "A2"

        if businesses:
            # Infer columns from keys of first row
            all_keys = list(
                dict.fromkeys(k for b in businesses for k in b.keys())
            )
            # Preferred order
            preferred = [
                "company", "website", "email_hints", "linkedin",
                "description", "phone", "twitter", "instagram", "facebook",
            ]
            ordered = [k for k in preferred if k in all_keys]
            ordered += [k for k in all_keys if k not in ordered]

            # Human-readable headers
            _label = {
                "company": "Company", "website": "Website",
                "email_hints": "Email", "linkedin": "LinkedIn",
                "description": "Description / Notes",
                "phone": "Phone", "twitter": "Twitter/X",
                "instagram": "Instagram", "facebook": "Facebook",
            }
            headers = ["#"] + [_label.get(k, k.replace("_", " ").title()) for k in ordered]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = _hdr_font
                cell.fill = _hdr_fill
                cell.alignment = _ctr
                cell.border = _border

            for i, b in enumerate(businesses, 1):
                row = [i]
                for k in ordered:
                    v = b.get(k, "")
                    if isinstance(v, list):
                        v = ", ".join(str(x) for x in v) if v else "NOT FOUND"
                    row.append(str(v)[:300] if v else "NOT FOUND")
                ws.append(row)
                ri = i + 1
                if i % 2 == 0:
                    for cell in ws[ri]:
                        cell.fill = _alt_fill
                for cell in ws[ri]:
                    cell.alignment = _wrap
                    cell.border = _border

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        else:
            # No structured data — write summary as table rows
            ws.append(["Output"])
            ws["A1"].font = _hdr_font
            ws["A1"].fill = _hdr_fill
            ws.column_dimensions["A"].width = 120
            for line in summary.split("\n"):
                ws.append([line])

        # ── Sheet 2: Full Report ────────────────────────────────────
        ws_rpt = wb.create_sheet("Full Report")
        ws_rpt.column_dimensions["A"].width = 120
        ws_rpt.append([f"Goal: {goal[:200]}"])
        ws_rpt["A1"].font = Font(bold=True, size=12)
        ws_rpt.append([""])
        for line in summary.split("\n"):
            ws_rpt.append([line])

        # ── Sheet 3: Outreach (if present) ─────────────────────────
        outreach_lines = self._extract_section(summary, "PHASE 3", "OUTREACH")
        if outreach_lines:
            ws_out = wb.create_sheet("Outreach Messages")
            ws_out.column_dimensions["A"].width = 110
            ws_out.append(["Outreach Messages"])
            ws_out["A1"].font = Font(bold=True, size=13)
            ws_out.append([""])
            for line in outreach_lines:
                ws_out.append([line])

        wb.save(str(path))
        return str(path)

    # ── JSON ────────────────────────────────────────────────────────
    def _export_json(self, goal: str, result: dict) -> Optional[str]:
        ts   = _ts()
        path = _desktop() / f"MegaV_{_slug(goal)}_{ts}.json"

        payload: dict[str, Any] = {
            "goal":      goal[:300],
            "exported":  datetime.datetime.now().isoformat(),
            "success":   result.get("success", False),
        }
        for key in ("businesses", "leads", "data", "results", "records"):
            if result.get(key):
                payload[key] = result[key]
                break

        # Embed summary as text if no structured data
        if not any(k in payload for k in ("businesses", "leads", "data", "results", "records")):
            payload["output"] = result.get("summary") or result.get("content") or ""

        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return str(path)

    # ── PDF ─────────────────────────────────────────────────────────
    def _export_pdf(self, goal: str, result: dict) -> Optional[str]:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib import colors
        except ImportError:
            if not _pip("reportlab"):
                return self._export_txt(goal, result)
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib import colors

        ts   = _ts()
        path = _desktop() / f"MegaV_{_slug(goal)}_{ts}.pdf"
        doc  = SimpleDocTemplate(str(path), pagesize=A4,
                                  leftMargin=20*mm, rightMargin=20*mm,
                                  topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        h1 = ParagraphStyle("H1", parent=styles["Heading1"],
                             textColor=colors.HexColor("#1F4E79"), spaceAfter=8)
        body = ParagraphStyle("Body", parent=styles["Normal"],
                               fontSize=9, leading=13, spaceAfter=4)

        story = []
        story.append(Paragraph("MegaV Task Report", h1))
        story.append(Paragraph(f"Goal: {goal[:200]}", styles["Italic"]))
        story.append(Paragraph(
            f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Italic"],
        ))
        story.append(Spacer(1, 12))

        summary = result.get("summary") or result.get("content") or ""
        for line in summary.split("\n"):
            txt = line.strip()
            if not txt:
                story.append(Spacer(1, 4))
                continue
            escaped = txt.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            if txt.startswith("=") and len(txt) > 20:
                story.append(Paragraph(escaped, h1))
            else:
                story.append(Paragraph(escaped, body))

        doc.build(story)
        return str(path)

    # ── HTML dashboard ─────────────────────────────────────────────
    def _export_html(self, goal: str, result: dict) -> Optional[str]:
        ts       = _ts()
        path     = _desktop() / f"MegaV_{_slug(goal)}_{ts}.html"
        summary  = result.get("summary") or result.get("content") or ""
        businesses: list[dict] = result.get("businesses") or []

        # Build table rows if we have structured data
        table_html = ""
        if businesses:
            hdrs = list(dict.fromkeys(k for b in businesses for k in b.keys()))
            thead = "".join(f"<th>{h.replace('_',' ').title()}</th>" for h in hdrs)
            rows_html = ""
            for i, b in enumerate(businesses):
                tds = ""
                for k in hdrs:
                    v = b.get(k, "")
                    if isinstance(v, list):
                        v = ", ".join(str(x) for x in v) or "NOT FOUND"
                    v = str(v)[:200] if v else "NOT FOUND"
                    if v.startswith("http"):
                        v = f'<a href="{v}" target="_blank">{v}</a>'
                    tds += f"<td>{v}</td>"
                cls = "even" if i % 2 == 0 else ""
                rows_html += f'<tr class="{cls}">{tds}</tr>'
            table_html = f"""
            <h2>Lead Table</h2>
            <table><thead><tr>{thead}</tr></thead><tbody>{rows_html}</tbody></table>
            """

        # Format plain summary as paragraphs
        paras = ""
        for line in summary.split("\n"):
            t = line.strip()
            if not t:
                paras += "<br>"
            elif t.startswith("="):
                paras += f"<h3>{t.strip('=').strip()}</h3>"
            else:
                paras += f"<p>{t}</p>"

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>MegaV — {goal[:60]}</title>
<style>
  body{{font-family:Segoe UI,Arial,sans-serif;background:#0d1117;color:#cdd9e5;margin:0;padding:20px}}
  h1{{color:#4d9fff;border-bottom:2px solid #1f4e79;padding-bottom:8px}}
  h2{{color:#79c0ff;margin-top:28px}}
  h3{{color:#3fb950;border-left:3px solid #3fb950;padding-left:10px}}
  p{{line-height:1.6;margin:4px 0}}
  table{{width:100%;border-collapse:collapse;margin-top:12px;font-size:13px}}
  th{{background:#1f4e79;color:#fff;padding:8px 10px;text-align:left}}
  td{{padding:7px 10px;border-bottom:1px solid #21262d}}
  tr.even td{{background:#161b22}}
  a{{color:#58a6ff}}
  .meta{{color:#8b949e;font-size:12px;margin-bottom:16px}}
</style>
</head>
<body>
<h1>MegaV Task Report</h1>
<p class="meta">Goal: {goal[:200]} &nbsp;|&nbsp; {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
{table_html}
<h2>Full Report</h2>
{paras}
</body>
</html>"""

        path.write_text(html, encoding="utf-8")
        return str(path)

    # ── Code file ────────────────────────────────────────────────────
    def _export_code(self, goal: str, result: dict) -> Optional[str]:
        summary = result.get("summary") or result.get("content") or ""
        # Extract first code block
        m = re.search(r"```(?:python|py|bash|sh|js|javascript)?\n(.*?)```",
                      summary, re.DOTALL)
        code = m.group(1).strip() if m else summary

        # Detect language
        ext = ".py"
        if re.search(r"```(?:js|javascript)", summary, re.IGNORECASE):
            ext = ".js"
        elif re.search(r"```(?:bash|sh)", summary, re.IGNORECASE):
            ext = ".sh"

        ts   = _ts()
        path = _desktop() / f"MegaV_{_slug(goal)}_{ts}{ext}"
        path.write_text(code, encoding="utf-8")
        return str(path)

    # ── ZIP package ─────────────────────────────────────────────────
    def _export_zip(self, goal: str, result: dict) -> Optional[str]:
        ts       = _ts()
        zip_path = _desktop() / f"MegaV_{_slug(goal)}_{ts}.zip"
        summary  = result.get("summary") or result.get("content") or ""

        with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
            # Always include full text report
            zf.writestr("report.txt", summary)

            # Include JSON if structured data present
            if result.get("businesses") or result.get("leads"):
                data = result.get("businesses") or result.get("leads")
                zf.writestr("data.json", json.dumps(data, indent=2))

            # Include any extra files referenced in result
            for key in ("path", "file", "output_file"):
                p = result.get(key)
                if p and Path(p).exists():
                    zf.write(p, Path(p).name)

        return str(zip_path)

    # ── Plain text ───────────────────────────────────────────────────
    def _export_txt(self, goal: str, result: dict) -> Optional[str]:
        ts       = _ts()
        path     = _desktop() / f"MegaV_{_slug(goal)}_{ts}.txt"
        summary  = result.get("summary") or result.get("content") or ""
        goal_pre = goal[:200] + ("…" if len(goal) > 200 else "")
        header   = (
            "MegaV Task Output\n"
            + "=" * 60 + "\n"
            + f"Goal : {goal_pre}\n"
            + f"Date : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            + "=" * 60 + "\n\n"
        )
        path.write_text(header + summary, encoding="utf-8")
        return str(path)

    # ------------------------------------------------------------------
    # Per-phase export (for compound multi-phase goals)
    # ------------------------------------------------------------------

    def export_phase(
        self,
        goal: str,
        phase_idx: int,
        phase_name: str,
        action: str,
        result: dict,
    ) -> str:
        """Save a single phase's output to a file on the Desktop.

        Returns the saved file path, or "" if the output was too small to save.
        """
        summary = result.get("summary") or result.get("content") or ""
        if len(summary.strip()) < 150:
            return ""

        slug = _slug(goal, 28)
        phase_slug = re.sub(r"[^\w]+", "_", phase_name[:28]).strip("_")
        fname_base = f"MegaV_{slug}_Phase{phase_idx:02d}_{phase_slug}"

        fmt = self._detect_phase_format(action, summary)
        try:
            return self._write_phase_file(fname_base, fmt, summary, phase_name) or ""
        except Exception:
            try:
                p = _desktop() / f"{fname_base}.txt"
                p.write_text(summary, encoding="utf-8")
                return str(p)
            except Exception:
                return ""

    def _detect_phase_format(self, action: str, summary: str) -> str:
        """Pick the best file format for a phase based on its action and content."""
        # generate_code phases: check for HTML first, then Python
        if action == "generate_code":
            if re.search(r"<!doctype html|<html|<body", summary, re.IGNORECASE):
                return "html"
            if re.search(r"```html", summary, re.IGNORECASE):
                return "html"
            if re.search(r"```(?:python|py)\n", summary, re.IGNORECASE):
                return "code"
            return "html"  # generate_code default

        # Content clearly contains HTML
        if re.search(r"<!doctype html|<html\b", summary, re.IGNORECASE):
            return "html"
        if re.search(r"```html", summary, re.IGNORECASE):
            return "html"

        # Python/JS code blocks
        if re.search(r"```(?:python|py|javascript|js|bash|sh)\n", summary, re.IGNORECASE):
            return "code"

        return "txt"

    def _write_phase_file(self, fname_base: str, fmt: str, summary: str, phase_name: str) -> str:
        """Write phase content to disk in the chosen format. Returns file path."""
        if fmt == "html":
            # Try to extract a clean HTML block from the summary
            # 1. Fenced code block: ```html ... ```
            m = re.search(r"```html\n(.*?)```", summary, re.IGNORECASE | re.DOTALL)
            if m:
                html_content = m.group(1).strip()
            # 2. Raw <!DOCTYPE … </html>
            elif re.search(r"<!doctype html", summary, re.IGNORECASE):
                m2 = re.search(r"(<!DOCTYPE html.*?</html>)", summary, re.IGNORECASE | re.DOTALL)
                html_content = m2.group(1).strip() if m2 else summary
            else:
                # Wrap plain text as minimal HTML
                lines = "".join(
                    f"<h2>{l.strip('=# ').strip()}</h2>" if l.strip().startswith(("=", "#"))
                    else f"<p>{l}</p>"
                    for l in summary.split("\n")
                )
                html_content = (
                    f'<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">'
                    f'<title>{phase_name}</title>'
                    f'<style>body{{font-family:Segoe UI,sans-serif;max-width:900px;'
                    f'margin:40px auto;line-height:1.7;color:#222}}'
                    f'h2{{color:#1f4e79}}</style></head><body>{lines}</body></html>'
                )
            path = _desktop() / f"{fname_base}.html"
            path.write_text(html_content, encoding="utf-8")
            return str(path)

        if fmt == "code":
            m = re.search(
                r"```(?:python|py|javascript|js|bash|sh)?\n(.*?)```",
                summary, re.DOTALL,
            )
            code = m.group(1).strip() if m else summary
            ext = ".js" if re.search(r"```(?:js|javascript)", summary, re.IGNORECASE) else ".py"
            path = _desktop() / f"{fname_base}{ext}"
            path.write_text(code, encoding="utf-8")
            return str(path)

        # txt default
        path = _desktop() / f"{fname_base}.txt"
        path.write_text(summary, encoding="utf-8")
        return str(path)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_section(text: str, *markers: str) -> list[str]:
        """Extract lines between the first matching section header and the next '='*30 divider."""
        lines = text.split("\n")
        collecting = False
        out: list[str] = []
        for line in lines:
            if all(m.upper() in line.upper() for m in markers):
                collecting = True
                continue
            if collecting:
                if line.startswith("=" * 30):
                    break
                out.append(line)
        return out


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_engine: Optional[OutputEngine] = None


def get_output_engine() -> OutputEngine:
    global _engine
    if _engine is None:
        _engine = OutputEngine()
    return _engine
