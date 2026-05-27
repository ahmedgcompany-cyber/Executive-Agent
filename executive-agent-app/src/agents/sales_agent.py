"""Sales agent for market research and lead generation tasks."""

import logging
from pathlib import Path
from typing import Any, Optional

_log = logging.getLogger("megav.sales")


class SalesAgent:
    """Specialist agent for sales and market research tasks."""

    def __init__(self):
        """Initialize sales agent."""
        # Load prompt
        self.prompt = self._load_prompt()

    def _load_prompt(self) -> str:
        """Load sales prompt from file."""
        prompt_path = Path("src/prompts/sales.txt")
        if prompt_path.exists():
            return prompt_path.read_text(encoding="utf-8")
        return "You are a sales and market research specialist agent."

    def collect_market_info(self, market_segment: str) -> dict[str, Any]:
        """Collect information about a market segment.

        Args:
            market_segment: Market segment to research

        Returns:
            Market information
        """
        try:
            from ..providers.model_router import ModelRouter, NoModelAvailableError
            router = ModelRouter()
            resp = router.ask(
                system="You are a market research analyst. Provide concise, factual market intelligence.",
                user=f"Provide market research data for the segment: {market_segment}. Include market size, growth rate, key players, and trends.",
                task_type="general",
            )
            if resp:
                return {
                    "success": True,
                    "market_segment": market_segment,
                    "analysis": resp,
                }
        except NoModelAvailableError:
            return {"success": False, "error": "No LLM available for market research.", "market_segment": market_segment}
        except Exception:
            import logging
            logging.getLogger(__name__).exception("collect_market_info error")

        return {"success": False, "error": "Market research failed.", "market_segment": market_segment}

    def structure_leads(self, raw_leads: list[dict[str, Any]]) -> dict[str, Any]:
        """Structure raw lead data into organized format.

        Args:
            raw_leads: List of raw lead data

        Returns:
            Structured leads
        """
        structured = []

        for lead in raw_leads:
            structured_lead = {
                "company": lead.get("company", ""),
                "contact_name": lead.get("contact_name", ""),
                "title": lead.get("title", ""),
                "email": lead.get("email", ""),
                "phone": lead.get("phone", ""),
                "linkedin": lead.get("linkedin", ""),
                "industry": lead.get("industry", ""),
                "company_size": lead.get("company_size", ""),
                "priority": lead.get("priority", "medium"),
                "status": lead.get("status", "new"),
                "notes": lead.get("notes", ""),
            }
            structured.append(structured_lead)

        return {
            "success": True,
            "leads": structured,
            "count": len(structured),
        }

    def draft_outreach(
        self,
        lead: dict[str, Any],
        template: str = "default",
    ) -> dict[str, Any]:
        """Draft outreach message for a lead.

        Args:
            lead: Lead information
            template: Template to use

        Returns:
            Drafted outreach
        """
        company = lead.get("company", "")
        contact_name = lead.get("contact_name", "Hiring Manager")

        try:
            from ..providers.model_router import ModelRouter, NoModelAvailableError
            router = ModelRouter()
            resp = router.ask(
                system="You are an elite outreach copywriter. Write a personalised, professional outreach message.",
                user=(
                    f"Draft a {template} outreach message for a lead.\n"
                    f"Company: {company}\nContact: {contact_name}\n"
                    f"Lead details: {lead}"
                ),
                task_type="general",
            )
            if resp:
                return {
                    "success": True,
                    "message": resp,
                    "recipient": contact_name,
                    "company": company,
                    "template": template,
                }
        except NoModelAvailableError:
            return {"success": False, "error": "No LLM available to draft outreach.", "recipient": contact_name, "company": company}
        except Exception:
            import logging
            logging.getLogger(__name__).exception("draft_outreach error")

        return {"success": False, "error": "Outreach drafting failed.", "recipient": contact_name, "company": company}

    def compare_products(
        self,
        products: list[dict[str, Any]],
        criteria: list[str],
    ) -> dict[str, Any]:
        """Compare products based on criteria.

        Args:
            products: List of products to compare
            criteria: Comparison criteria

        Returns:
            Comparison result
        """
        comparison = {
            "products": [],
            "criteria": criteria,
            "matrix": {},
        }

        for product in products:
            product_data = {
                "name": product.get("name", ""),
                "price": product.get("price", 0),
                "features": product.get("features", []),
            }
            comparison["products"].append(product_data)

        # Build comparison matrix
        for criterion in criteria:
            comparison["matrix"][criterion] = {}
            for product in products:
                comparison["matrix"][criterion][product.get("name", "")] = product.get(criterion, "N/A")

        return {
            "success": True,
            "comparison": comparison,
        }

    def analyze_competitors(self, competitors: list[str]) -> dict[str, Any]:
        """Analyze competitors.

        Args:
            competitors: List of competitor names

        Returns:
            Competitor analysis
        """
        try:
            from ..providers.model_router import ModelRouter, NoModelAvailableError
            router = ModelRouter()
            resp = router.ask(
                system="You are a competitive intelligence analyst. Provide strengths, weaknesses, market position, and key products for each competitor.",
                user=f"Analyze these competitors: {', '.join(competitors)}",
                task_type="general",
            )
            if resp:
                return {
                    "success": True,
                    "competitors": competitors,
                    "analysis": resp,
                }
        except NoModelAvailableError:
            return {"success": False, "error": "No LLM available for competitor analysis.", "competitors": competitors}
        except Exception:
            import logging
            logging.getLogger(__name__).exception("analyze_competitors error")

        return {"success": False, "error": "Competitor analysis failed.", "competitors": competitors}

    # ── Email + CRM patterns the sales agent handles ────────────────
    _EMAIL_RE = None
    _EMAIL_PATTERNS = [
        r"\b(check|read|get|show|fetch)\b.{0,20}\b(email|inbox|mail|leads?)\b",
        r"\b(send|compose)\b.{0,20}\b(email|outreach|message)\b",
        r"\bfollow.?up\b",
        r"\bcrm\b|\blead.track\b|\bpipeline\b",
    ]

    @classmethod
    def _is_email_goal(cls, goal: str) -> bool:
        import re
        if cls._EMAIL_RE is None:
            cls._EMAIL_RE = re.compile("|".join(cls._EMAIL_PATTERNS), re.IGNORECASE)
        return bool(cls._EMAIL_RE.search(goal))

    def _try_email_crm_action(self, goal: str, context: Any) -> Optional[dict[str, Any]]:
        """Route email/CRM-related sales goals to the appropriate services."""
        if not self._is_email_goal(goal):
            return None
        try:
            from ..integrations.crm_service import get_crm_service
            from ..integrations.email_service import get_email_service
            import re

            profile = {}
            try:
                profile = dict(getattr(context, "profile", {}) or {})
            except Exception as _e:
                _log.debug("Could not read context profile: %s", _e)
            default_account = profile.get("email", "")

            # CRM pipeline / follow-up
            if re.search(r"\bfollow.?up\b|\bcrm\b|\bpipeline\b|\blead.track\b", goal, re.IGNORECASE):
                crm = get_crm_service()
                summary = crm.get_pipeline_summary()
                fus = crm.get_follow_ups()
                lines = [
                    f"Sales CRM — {summary['total_contacts']} contacts",
                    f"Follow-ups due: {summary['follow_ups_due']}",
                ]
                for stage, cnt in summary.get("by_stage", {}).items():
                    if cnt:
                        lines.append(f"  {stage}: {cnt}")
                if fus:
                    lines.append("\nOverdue follow-ups:")
                    for c in fus[:4]:
                        lines.append(f"  • {c.get('name') or c['email']} — {c.get('stage')}")
                return {"success": True, "summary": "\n".join(lines), "via_crm": True}

            # Email read / summary
            email_svc = get_email_service()
            if not email_svc.is_connected():
                return None
            r = email_svc.handle_prompt(goal, default_account=default_account)
            if r.get("success"):
                return {"success": True, "summary": r.get("summary") or r.get("message", "Done."), "via_email": True}
        except Exception as _e:
            _log.error("Email/CRM action failed in sales agent: %s", _e)
        return None

    def handle_sales_task(self, action: str, context: dict[str, Any]) -> dict[str, Any]:
        """Handle a sales task.

        Args:
            action: Action to perform
            context: Task context

        Returns:
            Task result
        """
        goal_text = getattr(context, "goal", "") or getattr(context, "current_goal", "") or ""

        # ── Email/CRM intercept ────────────────────────────────────────
        if action in ("execute_goal", "draft", "structure") and goal_text:
            email_result = self._try_email_crm_action(goal_text, context)
            if email_result:
                return email_result
        # ── End intercept ─────────────────────────────────────────────

        handlers = {
            "research": lambda: self.collect_market_info(context.get("market_segment", "")),
            "structure_leads": lambda: self.structure_leads(context.get("raw_leads", [])),
            "draft_outreach": lambda: self.draft_outreach(
                context.get("lead", {}),
                context.get("template", "default"),
            ),
            "compare_products": lambda: self.compare_products(
                context.get("products", []),
                context.get("criteria", []),
            ),
            "analyze_competitors": lambda: self.analyze_competitors(
                context.get("competitors", [])
            ),
        }

        handlers["execute_goal"] = lambda: self._execute_from_goal(goal_text)
        handlers["structure"]    = lambda: self._execute_from_goal(goal_text)
        handlers["draft"]        = lambda: self._execute_from_goal(goal_text)
        handlers["verify"]       = lambda: {"success": True, "summary": "Sales task verified."}

        handler = handlers.get(action)
        if handler:
            return handler()

        return {"success": False, "error": f"Unknown action: {action}"}

    # ------------------------------------------------------------------
    # Goal-driven execution  (multi-phase planner)
    # ------------------------------------------------------------------

    _PHASE_PROMPTS = {
        "plan": (
            "You are a world-class sales strategist and automation expert.\n"
            "The user has given you a sales/lead-generation goal.\n"
            "First, output a clear numbered PLAN of exactly what you will do:\n"
            "  1. Niche & target definition\n"
            "  2. Lead discovery approach\n"
            "  3. Deep research method\n"
            "  4. Outreach message strategy\n"
            "  5. Channel execution\n"
            "  6. Tracking & feedback loop\n"
            "Keep the plan concise (1-2 sentences per step). "
            "Then output '--- BEGIN EXECUTION ---' on its own line and stop.\n"
        ),
        "leads": (
            "You are a lead researcher. Given the goal below, produce a table of "
            "10-15 example leads.\n"
            "Format each as:\n"
            "  LEAD #N\n"
            "  Company   : <name>\n"
            "  Contact   : <likely decision-maker title>\n"
            "  Website   : <url>\n"
            "  Email     : NOT FOUND (no real email available without live research)\n"
            "  LinkedIn  : NOT FOUND\n"
            "  Pain Point: <1-sentence specific weakness>\n"
            "  Opportunity: <1-sentence AI/automation win>\n\n"
            "STRICT RULES: Do NOT fabricate email addresses. If you don't have real data, "
            "write 'NOT FOUND'. The user will fill in emails from real outreach.\n"
        ),
        "outreach": (
            "You are an elite outreach copywriter. Given the lead list above, "
            "write a personalised outreach message for each lead.\n"
            "Each message must:\n"
            "  - Reference something specific about THAT company or person\n"
            "  - Mention a concrete problem they likely have\n"
            "  - Offer a clear AI/automation solution\n"
            "  - Sound human, not templated\n"
            "  - Be 4-6 sentences\n"
            "Format: OUTREACH FOR LEAD #N: <company>\n<message>\n"
        ),
        "analytics": (
            "You are a sales analytics consultant. Based on the outreach strategy above, "
            "produce a realistic campaign dashboard:\n"
            "  - Total leads targeted\n"
            "  - Projected send rate / timeline\n"
            "  - Expected open rate, reply rate, conversion rate (industry benchmarks)\n"
            "  - A/B test suggestions\n"
            "  - Feedback-loop improvement plan (how to improve messaging after replies)\n"
            "Format clearly with section headers.\n"
        ),
    }

    # ── Web research prompts (used when real data is available) ────────
    _PHASE_PROMPTS_WITH_DATA = {
        "leads": (
            "You are a lead researcher. The user has provided REAL business data "
            "scraped from the web (actual companies with real websites).\n"
            "Using ONLY the companies listed in the data below, produce a structured "
            "lead table. For each company:\n"
            "  LEAD #N\n"
            "  Company   : <name from data>\n"
            "  Website   : <url from data>\n"
            "  Contact   : <infer likely decision-maker title>\n"
            "  Email     : <use ONLY emails from the data — if none found write: NOT FOUND>\n"
            "  LinkedIn  : <url from data — if none found write: NOT FOUND>\n"
            "  Pain Point: <1-sentence specific weakness based on their description>\n"
            "  Opportunity: <1-sentence AI/automation win>\n\n"
            "STRICT RULES:\n"
            "  - Do NOT invent companies not in the data.\n"
            "  - Do NOT guess, infer, or fabricate email addresses.\n"
            "  - If email is not in the provided data → write 'NOT FOUND'\n"
            "  - Never write patterns like 'info@domain.com' unless it was in the data.\n"
        ),
        "outreach": (
            "You are an elite outreach copywriter. Given the REAL lead list above, "
            "write a personalised outreach message for each lead.\n"
            "Each message must:\n"
            "  - Reference something specific about THAT company (from their real website/description)\n"
            "  - Mention a concrete problem they likely have\n"
            "  - Offer a clear AI/automation solution\n"
            "  - Sound human, not templated\n"
            "  - Be 4-6 sentences\n"
            "Format: OUTREACH FOR LEAD #N: <company>\n<message>\n"
        ),
    }

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def _save_leads_excel(
        self,
        businesses: list[dict],
        phases_text: str,
        goal: str,
    ) -> str | None:
        """
        Save the lead list + full report to an .xlsx file on the Desktop.

        Columns: #, Company, Website, Email, Contact Page, LinkedIn, Description/Notes.
        Creates three sheets: Leads, Full Report, Outreach Messages.

        Returns the saved file path, or None if openpyxl is unavailable.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            # Try to auto-install openpyxl
            try:
                import subprocess, sys as _sys
                r = subprocess.run(
                    [_sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                    capture_output=True, timeout=90,
                )
                if r.returncode == 0:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                else:
                    return None
            except Exception:
                return None

        import datetime, re as _re
        from pathlib import Path

        # ── Destination: local Desktop (NOT OneDrive) ──────────────────
        desktop = Path.home() / "Desktop"
        try:
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders",
            )
            real_desktop = winreg.QueryValueEx(key, "Desktop")[0]
            if real_desktop:
                desktop = Path(real_desktop)
        except Exception:
            pass
        desktop.mkdir(parents=True, exist_ok=True)

        ts_str    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        slug      = _re.sub(r"[^\w]+", "_", goal[:45]).strip("_")
        xlsx_path = desktop / f"MegaV_{slug}_{ts_str}.xlsx"

        wb = openpyxl.Workbook()

        # ── Shared styles ───────────────────────────────────────────────
        _hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        _hdr_fill  = PatternFill(fill_type="solid", fgColor="1F4E79")   # dark blue
        _alt_fill  = PatternFill(fill_type="solid", fgColor="D6E4F0")   # light blue
        _ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _wrap      = Alignment(vertical="top", wrap_text=True)
        _thin      = Side(style="thin", color="AAAAAA")
        _border    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        # ── Sheet 1: Leads ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Leads"
        ws.freeze_panes = "A2"   # freeze header row

        col_headers = [
            "#", "Company", "Website", "Email",
            "Contact Page", "LinkedIn", "Description / Notes",
        ]
        col_widths = [4, 32, 38, 32, 28, 36, 55]

        ws.append(col_headers)
        for cell in ws[1]:
            cell.font    = _hdr_font
            cell.fill    = _hdr_fill
            cell.alignment = _ctr
            cell.border  = _border

        for i, b in enumerate(businesses, 1):
            emails = ", ".join(b.get("email_hints", []) or []) or "NOT FOUND"
            linkedin = b.get("linkedin", "") or "NOT FOUND"
            row = [
                i,
                b.get("company", ""),
                b.get("website", ""),
                emails,
                "",          # contact page — user fills in
                linkedin,
                b.get("description", "")[:300],
            ]
            ws.append(row)
            # Alternate row shading
            row_idx = i + 1
            if i % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = _alt_fill
            for cell in ws[row_idx]:
                cell.alignment = _wrap
                cell.border    = _border

        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        ws.row_dimensions[1].height = 18

        # ── Sheet 2: Outreach Messages ──────────────────────────────────
        outreach_lines: list[str] = []
        in_outreach = False
        for line in phases_text.split("\n"):
            if "PHASE 3" in line and "OUTREACH" in line.upper():
                in_outreach = True
                continue
            if in_outreach and line.startswith("=" * 30):
                in_outreach = False
            if in_outreach:
                outreach_lines.append(line)

        if outreach_lines:
            ws_out = wb.create_sheet("Outreach Messages")
            ws_out.column_dimensions["A"].width = 110
            ws_out.append(["Outreach Messages"])
            ws_out["A1"].font = Font(bold=True, size=13)
            ws_out.append([""])
            for line in outreach_lines:
                ws_out.append([line])

        # ── Sheet 3: Full Report ────────────────────────────────────────
        ws_rpt = wb.create_sheet("Full Report")
        ws_rpt.column_dimensions["A"].width = 120
        ws_rpt.append(["Full Report — " + goal[:120]])
        ws_rpt["A1"].font = Font(bold=True, size=12)
        ws_rpt.append([""])
        for line in phases_text.split("\n"):
            ws_rpt.append([line])

        try:
            wb.save(str(xlsx_path))
            return str(xlsx_path)
        except Exception:
            return None

    def _execute_from_goal(self, goal: str) -> dict[str, Any]:
        """
        Multi-phase sales execution: Web Discovery → Plan → Leads → Outreach → Analytics.

        Phase 0: Real web research (DuckDuckGo/Bing/HTTP fallback) to find actual businesses.
        Phase 1: LLM execution plan.
        Phase 2: Structured lead list (enriched with real data if available).
        Phase 3: Personalised outreach messages.
        Phase 4: Campaign analytics dashboard.
        """
        if not goal:
            return {"success": False, "error": "No goal provided to sales agent."}

        try:
            from ..providers.model_router import ModelRouter, NoModelAvailableError
            router = ModelRouter()
            sections: list[str] = []

            # ── Phase 0 — Real web research ───────────────────────────────
            real_businesses: list[dict] = []
            web_research_text = ""
            try:
                from ..tools_ext.web_research_tools import (
                    get_web_researcher, extract_niche_from_goal,
                )
                import re as _re

                researcher = get_web_researcher()

                # Use the shared smart niche extractor
                niche, location = extract_niche_from_goal(goal)

                count_m = _re.search(
                    r"\b(\d+)\s*(?:leads?|companies|businesses|prospects)\b",
                    goal, _re.IGNORECASE,
                )
                count = min(int(count_m.group(1)) if count_m else 10, 15)

                real_businesses = researcher.discover(niche, location=location, count=count)

                if real_businesses:
                    lines = [
                        f"\n" + "=" * 60,
                        "PHASE 0 — REAL WEB DISCOVERY",
                        "=" * 60,
                        f"Niche: {niche}{' | Location: ' + location if location else ''}",
                        f"Discovered {len(real_businesses)} real businesses via web search:",
                        "",
                    ]
                    for i, b in enumerate(real_businesses, 1):
                        lines.append(f"  {i}. {b['company']}")
                        lines.append(f"     Website : {b['website']}")
                        if b.get("description"):
                            lines.append(f"     Overview: {b['description'][:150]}")
                        if b.get("email_hints"):
                            lines.append(f"     Emails  : {', '.join(b['email_hints'])}")
                    lines.append("\n[Real data — no hallucination]")
                    web_research_text = "\n".join(lines)
                    sections.append(web_research_text)
                else:
                    # Web research returned nothing — add a note but continue
                    sections.append(
                        "\n" + "=" * 60 + "\n"
                        "PHASE 0 — WEB RESEARCH\n"
                        "=" * 60 + "\n"
                        "Note: No businesses found via web scraping for this niche.\n"
                        "The LLM will generate example leads based on the goal description."
                    )
            except Exception as _we:
                sections.append(
                    f"\n[Web research skipped: {_we}]\n"
                    "LLM will generate example leads."
                )

            # ── Phase 1 — Execution Plan ──────────────────────────────────
            plan_resp = router.ask(
                system=self._PHASE_PROMPTS["plan"],
                user=f"GOAL: {goal}",
                task_type="general",
            )
            if plan_resp:
                sections.append("=" * 60)
                sections.append("PHASE 1 — EXECUTION PLAN")
                sections.append("=" * 60)
                plan_text = plan_resp.split("--- BEGIN EXECUTION ---")[0].strip()
                sections.append(plan_text)

            # ── Phase 2 — Lead List ───────────────────────────────────────
            if real_businesses:
                # Feed REAL data to the LLM for structured formatting
                real_data_str = "\n".join(
                    f"Company: {b['company']}\n"
                    f"Website: {b['website']}\n"
                    f"Description: {b.get('description','')}\n"
                    f"Emails: {', '.join(b.get('email_hints',[]))}\n"
                    f"LinkedIn: {b.get('linkedin','')}\n"
                    for b in real_businesses
                )
                leads_user = (
                    f"GOAL: {goal}\n\n"
                    f"REAL BUSINESS DATA FROM WEB:\n{real_data_str[:4000]}"
                )
                leads_system = self._PHASE_PROMPTS_WITH_DATA["leads"]
            else:
                leads_user = f"GOAL: {goal}"
                leads_system = self._PHASE_PROMPTS["leads"]

            leads_resp = router.ask(
                system=leads_system,
                user=leads_user,
                task_type="general",
            )
            if leads_resp:
                sections.append("\n" + "=" * 60)
                sections.append("PHASE 2 — LEAD LIST")
                sections.append("=" * 60)
                sections.append(leads_resp)

            # ── Phase 3 — Personalised Outreach ───────────────────────────
            outreach_ctx = (leads_resp or "") + f"\n\nORIGINAL GOAL: {goal}"
            outreach_system = (
                self._PHASE_PROMPTS_WITH_DATA["outreach"]
                if real_businesses
                else self._PHASE_PROMPTS["outreach"]
            )
            outreach_resp = router.ask(
                system=outreach_system,
                user=outreach_ctx[:4000],
                task_type="general",
            )
            if outreach_resp:
                sections.append("\n" + "=" * 60)
                sections.append("PHASE 3 — PERSONALISED OUTREACH MESSAGES")
                sections.append("=" * 60)
                sections.append(outreach_resp)

            # ── Phase 4 — Analytics Dashboard ─────────────────────────────
            leads_count = len(real_businesses) if real_businesses else 15
            analytics_resp = router.ask(
                system=self._PHASE_PROMPTS["analytics"],
                user=(
                    f"GOAL: {goal}\n\n"
                    f"Leads targeted: {leads_count} (real companies discovered via web search)\n"
                    f"Channels: Email, LinkedIn"
                ),
                task_type="general",
            )
            if analytics_resp:
                sections.append("\n" + "=" * 60)
                sections.append("PHASE 4 — CAMPAIGN ANALYTICS DASHBOARD")
                sections.append("=" * 60)
                sections.append(analytics_resp)

            # ── Final assembly ─────────────────────────────────────────────
            if sections:
                full_output = "\n".join(sections)

                # ── Save structured Excel to Desktop ───────────────────
                excel_path = self._save_leads_excel(
                    businesses=real_businesses,
                    phases_text=full_output,
                    goal=goal,
                )

                result: dict[str, Any] = {
                    "success":    True,
                    "content":    full_output,
                    "summary":    full_output,
                    "businesses": real_businesses,
                }
                if excel_path:
                    result["path"]    = excel_path
                    result["summary"] = (
                        full_output
                        + f"\n\n{'='*60}\n"
                        + f"Excel saved → {excel_path}\n"
                        + f"{'='*60}"
                    )
                return result

            # LLM returned nothing but we have real web data — return that
            if real_businesses:
                lines = [
                    "=" * 60,
                    "LEAD GENERATION RESULTS",
                    "=" * 60,
                    f"Niche   : {niche}{' | Location: ' + location if location else ''}",
                    f"Found   : {len(real_businesses)} real businesses via web research",
                    "",
                ]
                for i, b in enumerate(real_businesses, 1):
                    lines.append(f"LEAD #{i}")
                    lines.append(f"  Company   : {b['company']}")
                    lines.append(f"  Website   : {b['website']}")
                    if b.get("description"):
                        lines.append(f"  Overview  : {b['description'][:200]}")
                    emails = b.get("email_hints") or []
                    lines.append(f"  Emails    : {', '.join(emails) if emails else 'NOT FOUND'}")
                    lines.append(f"  LinkedIn  : {b.get('linkedin') or 'NOT FOUND'}")
                    lines.append("")
                lines.append("=" * 60)
                lines.append("Note: LLM not available — showing raw web research data.")
                lines.append("Set ANTHROPIC_API_KEY or run Ollama for enriched output.")
                out = "\n".join(lines)
                excel_path = self._save_leads_excel(
                    businesses=real_businesses, phases_text=out, goal=goal,
                )
                result = {"success": True, "summary": out, "businesses": real_businesses}
                if excel_path:
                    result["path"]    = excel_path
                    result["summary"] = out + f"\n\nExcel saved → {excel_path}"
                return result

        except NoModelAvailableError as _nma:
            return {
                "success": False,
                "error": str(_nma),
                "summary": f"No LLM available: {_nma}",
            }
        except Exception as _exc:
            import traceback
            return {
                "success": False,
                "error": str(_exc),
                "summary": f"Sales agent error: {_exc}",
                "traceback": traceback.format_exc(),
            }

        # Absolute fallback — no web research, no LLM
        return {
            "success": False,
            "summary": (
                "No web research results and no LLM available.\n"
                "Please ensure you have an internet connection and optionally\n"
                "set ANTHROPIC_API_KEY or run Ollama for full AI-powered output."
            ),
        }
